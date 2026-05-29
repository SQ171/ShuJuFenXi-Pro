"""Excel 多 Sheet 导出"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
from analysis.pipeline import AnalysisContext
from config import ALL_METRICS


def export_to_excel(ctx: AnalysisContext, output_path: str):
    wb = Workbook()

    _sheet_force_eff(wb, ctx)
    _sheet_steady_state(wb, ctx)
    _sheet_degradation(wb, ctx)
    _sheet_anomaly(wb, ctx)
    _sheet_runtime(wb, ctx)

    wb.save(output_path)


def _header_style(cell, value: str):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    header_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    cell.value = value
    cell.font = Font(bold=True, size=11)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border


def _data_cell(cell, value):
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    cell.value = value
    cell.alignment = Alignment(horizontal="center")
    cell.border = thin_border


def _sheet_force_eff(wb: Workbook, ctx: AnalysisContext):
    ws = wb.active
    ws.title = "力效汇总"

    columns = [
        "文件名", "名义油门%", "循环序号", "力效均值(g/W)", "力效标准差",
        "力效CV%", "力效最小值", "力效最大值",
    ]
    for col_idx, col_name in enumerate(columns, 1):
        _header_style(ws.cell(row=1, column=col_idx), col_name)

    if ctx.step_summary is not None and not ctx.step_summary.empty:
        ss = ctx.step_summary
        row_idx = 2
        for _, row in ss.iterrows():
            if "force_eff_mean" not in row:
                continue
            vals = [
                row.get("source_file", ""),
                row.get("nominal_throttle", ""),
                row.get("cycle_num", ""),
                row.get("force_eff_mean", ""),
                row.get("force_eff_std", ""),
                row.get("force_eff_cv", ""),
                row.get("force_eff_min", ""),
                row.get("force_eff_max", ""),
            ]
            for col_idx, val in enumerate(vals, 1):
                _data_cell(ws.cell(row=row_idx, column=col_idx), val)
            row_idx += 1

    ws.freeze_panes = "A2"


def _sheet_steady_state(wb: Workbook, ctx: AnalysisContext):
    ws = wb.create_sheet("稳态段明细")
    if ctx.step_summary is None:
        return

    ss = ctx.step_summary
    columns = ["文件名", "循环序号", "名义油门%", "样本数"]
    for m in ALL_METRICS:
        columns.extend([
            f"{m.display_name}均值({m.unit})",
            f"{m.display_name}标准差",
            f"{m.display_name}CV%",
        ])

    for col_idx, col_name in enumerate(columns[:50], 1):
        _header_style(ws.cell(row=1, column=col_idx), col_name)

    row_idx = 2
    for _, row in ss.iterrows():
        vals = [
            row.get("source_file", ""),
            row.get("cycle_num", ""),
            row.get("nominal_throttle", ""),
            row.get("sample_count", ""),
        ]
        for m in ALL_METRICS:
            vals.append(row.get(f"{m.key}_mean", ""))
            vals.append(row.get(f"{m.key}_std", ""))
            vals.append(row.get(f"{m.key}_cv", ""))

        for col_idx, val in enumerate(vals[:50], 1):
            _data_cell(ws.cell(row=row_idx, column=col_idx), val)
        row_idx += 1

    ws.freeze_panes = "A2"


def _sheet_degradation(wb: Workbook, ctx: AnalysisContext):
    ws = wb.create_sheet("退化趋势")
    columns = ["指标", "维度", "分析类型", "油门%", "退化率/h", "R²", "p值", "显著", "样本数"]

    for col_idx, col_name in enumerate(columns, 1):
        _header_style(ws.cell(row=1, column=col_idx), col_name)

    row_idx = 2
    for r in ctx.degradation_report.results:
        vals = [r.metric_or_feature, r.dimension.value, r.analysis_type.value,
                r.nominal_throttle, round(r.slope_per_hour, 6), round(r.r_squared, 4),
                round(r.p_value, 4), "是" if r.is_significant else "否", r.data_points]
        for col_idx, val in enumerate(vals, 1):
            _data_cell(ws.cell(row=row_idx, column=col_idx), val)
        row_idx += 1

    ws.freeze_panes = "A2"


def _sheet_anomaly(wb: Workbook, ctx: AnalysisContext):
    ws = wb.create_sheet("异常数据")
    if ctx.anomaly_flags is None:
        return

    af = ctx.anomaly_flags
    outlier_count = af["is_outlier"].sum() if "is_outlier" in af.columns else 0
    ws.cell(row=1, column=1, value=f"异常数据总数: {outlier_count}")

    if outlier_count > 0:
        columns = ["is_outlier", "outlier_reason"]
        for col_idx, col_name in enumerate(columns, 1):
            _header_style(ws.cell(row=2, column=col_idx), col_name)

        row_idx = 3
        for _, row in af[af.get("is_outlier", False)].iterrows():
            for col_idx, col_name in enumerate(columns, 1):
                _data_cell(ws.cell(row=row_idx, column=col_idx), row.get(col_name, ""))
            row_idx += 1
            if row_idx > 1000:
                break


def _sheet_runtime(wb: Workbook, ctx: AnalysisContext):
    ws = wb.create_sheet("运行时间")
    if ctx.runtime_info is None:
        return

    columns = ["文件名", "运行时长(s)", "运行时长(min)"]
    for col_idx, col_name in enumerate(columns, 1):
        _header_style(ws.cell(row=1, column=col_idx), col_name)

    row_idx = 2
    for fname, secs in ctx.runtime_info.per_file_runtime.items():
        vals = [fname, round(secs, 1), round(secs / 60, 1)]
        for col_idx, val in enumerate(vals, 1):
            _data_cell(ws.cell(row=row_idx, column=col_idx), val)
        row_idx += 1

    r = row_idx + 1
    ws.cell(row=r, column=1, value=f"累计运行时间: {ctx.runtime_info.total_runtime_seconds/3600:.2f} h")
